import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, GradientFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from datetime import date

NAVY = '1B2A4A'
DARK_NAVY = '0D1B30'
GOLD = 'C9A84C'
WHITE = 'FFFFFF'
LIGHT_YELLOW = 'FFF9E6'
LIGHT_BLUE = 'E8F0FE'
SOFT_GRAY = 'F5F5F5'
RED_ALERT = 'FF4444'
YELLOW_WARN = 'FFD700'
GREEN_OK = '2ECC71'
PROTECTION_PASSWORD = 'smoney2025'

CATEGORIES = ['דיור', 'מזון', 'תחבורה', 'בריאות', 'פנאי ובידור',
              'קניות', 'חינוך', 'חיסכון והשקעות', 'מנויים', 'ילדים',
              'מתנות ותרומות', 'אחר']

SUBCATEGORIES = [
    'שכירות', 'משכנתא', 'חשמל', 'מים', 'גז', 'ועד בית', 'אינטרנט',
    'סופרמרקט', 'מסעדות', 'קפה', 'משלוחים', 'שווקים',
    'דלק', 'חניה', 'תחבורה ציבורית', 'אחזקת רכב', 'מונית/אובר',
    'ביטוח בריאות', 'רופא', 'תרופות', 'ספורט', 'מרפאה',
    'קולנוע', 'מנויים', 'חגים וטיולים', 'ספרים', 'בידור',
    'בגדים', 'נעליים', 'אלקטרוניקה', 'ריהוט',
    'קורסים', 'לימודים', 'חוגים',
    'קרן פנסיה', 'השקעות', 'חיסכון', 'קרן השתלמות',
    'נטפליקס', 'ספוטיפיי', 'אפליקציות',
    'גן/מעון', 'צעצועים', 'פעילויות',
]

def mfont(size=11, bold=False, color=NAVY, italic=False):
    return Font(name='Arial', size=size, bold=bold, color=color, italic=italic)

def mfill(color):
    return PatternFill('solid', fgColor=color)

def ralign(h='right', wrap=False, v='center'):
    return Alignment(horizontal=h, vertical=v, readingOrder=2, wrap_text=wrap)

def calign():
    return Alignment(horizontal='center', vertical='center', readingOrder=2)

def mborder(style='thin', color='CCCCCC'):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(border_style='medium', color=NAVY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(cell, text, size=12, fg=WHITE, bg=NAVY, bold=True, align='center'):
    cell.value = text
    cell.font = mfont(size=size, bold=bold, color=fg)
    cell.fill = mfill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', readingOrder=2)
    cell.border = mborder()

def input_cell(cell):
    cell.fill = mfill(LIGHT_YELLOW)
    cell.alignment = ralign()
    cell.border = mborder()
    cell.protection = Protection(locked=False)

def calc_cell(cell):
    cell.fill = mfill(LIGHT_BLUE)
    cell.alignment = calign()
    cell.border = mborder()

def set_defaults(ws, tab_color, zoom=100, freeze=None, grid=True, rtl=True):
    ws.sheet_properties.tabColor = tab_color
    if rtl:
        ws.sheet_view.rightToLeft = True
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.showGridLines = grid
    if freeze:
        ws.freeze_panes = freeze
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

def protect(ws, allow_sort=True, allow_filter=True):
    ws.protection.sheet = True
    ws.protection.password = PROTECTION_PASSWORD
    ws.protection.sort = not allow_sort
    ws.protection.autoFilter = not allow_filter
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


# ─── SHEET: נתונים (hidden data) ──────────────────────────────────────────────

def build_data_sheet(ws):
    set_defaults(ws, GOLD, grid=False)
    for i, cat in enumerate(CATEGORIES, 1):
        ws.cell(row=i, column=1, value=cat)
    for i, sub in enumerate(SUBCATEGORIES, 1):
        ws.cell(row=i, column=2, value=sub)
    income_sources = ['משכורת', 'עבודה עצמאית/פרילנס', 'בונוס', 'שכר דירה',
                      'דיבידנד/השקעות', 'מתנה', 'ירושה', 'קצבה', 'אחר']
    for i, s in enumerate(income_sources, 1):
        ws.cell(row=i, column=3, value=s)
    months = ['ינואר','פברואר','מרץ','אפריל','מאי','יוני',
              'יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר']
    for i, m in enumerate(months, 1):
        ws.cell(row=i, column=4, value=m)
    # Pie chart data (col E=category, col F=sumif amount)
    ws.cell(row=1, column=5, value='קטגוריה')
    ws.cell(row=1, column=6, value='סכום')
    for i, cat in enumerate(CATEGORIES, 2):
        ws.cell(row=i, column=5, value=cat)
        ws.cell(row=i, column=6,
                value=f"=IFERROR(SUMIF('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$B:$B,E{i},'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F:$F),0)")
    # Monthly bar chart data (col G=month, H=income, I=expenses)
    ws.cell(row=1, column=7, value='חודש')
    ws.cell(row=1, column=8, value='הכנסות')
    ws.cell(row=1, column=9, value='הוצאות')
    for m in range(1, 13):
        r = m + 1
        ws.cell(row=r, column=7, value=months[m-1])
        ws.cell(row=r, column=8,
                value=f"=IFERROR(SUMPRODUCT((MONTH('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$A$4:$A$503)={m})*(YEAR('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$A$4:$A$503)=2025)*'\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$E$4:$E$503),0)")
        ws.cell(row=r, column=9,
                value=f"=IFERROR(SUMPRODUCT((MONTH('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$A$4:$A$503)={m})*(YEAR('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$A$4:$A$503)=2025)*'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F$4:$F$503),0)")


# ─── SHEET: דף פתיחה (Cover) ──────────────────────────────────────────────────

def build_cover(ws):
    set_defaults(ws, GOLD, zoom=100, grid=False)
    ws.sheet_view.rightToLeft = True
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 9

    for r in range(1, 51):
        ws.row_dimensions[r].height = 28
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.fill = mfill(NAVY)
            cell.protection = Protection(locked=True)

    def mrow(r1, r2, c1, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        return ws.cell(r1, c1)

    c = mrow(1, 2, 1, 11)
    c.fill = mfill(DARK_NAVY)

    c = mrow(4, 8, 2, 10)
    c.value = 'לאן הכסף שלי נעלם?'
    c.font = mfont(size=36, bold=True, color=GOLD)
    c.alignment = calign()
    c.fill = mfill(NAVY)

    c = mrow(9, 10, 2, 10)
    c.value = 'מעקב פיננסי אישי — גרסה 2025'
    c.font = mfont(size=18, bold=False, color=WHITE, italic=True)
    c.alignment = calign()
    c.fill = mfill(NAVY)

    c = mrow(11, 11, 1, 11)
    c.fill = mfill(GOLD)

    c = mrow(13, 14, 2, 10)
    c.value = 'הכלי הפיננסי שישנה את הדרך שאתה מנהל את הכסף שלך'
    c.font = mfont(size=14, italic=True, color=WHITE)
    c.alignment = calign()
    c.fill = mfill(NAVY)

    bullets = [
        '✓   מעקב הכנסות והוצאות חודשי — פשוט ומהיר',
        '✓   לוח בקרה חכם עם גרפים אוטומטיים',
        '✓   ניתוח "לאן הכסף נעלם" — בקליק אחד',
        '✓   ניהול תקציב עם התראות אדום/צהוב/ירוק',
        '✓   סיכום שנתי מלא לכל 12 החודשים',
    ]
    for i, b in enumerate(bullets):
        row = 16 + i * 2
        c = mrow(row, row + 1, 2, 10)
        c.value = b
        c.font = mfont(size=13, bold=True, color=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        c.fill = mfill(NAVY)

    c = mrow(28, 28, 1, 11)
    c.fill = mfill(GOLD)

    c = mrow(30, 31, 2, 10)
    c.value = 'איך מתחילים?'
    c.font = mfont(size=16, bold=True, color=GOLD)
    c.alignment = calign()
    c.fill = mfill(DARK_NAVY)

    steps = ['1. עבור ללשונית "תקציב" והגדר יעדים חודשיים',
             '2. הזן הכנסות ב"הכנסות" והוצאות ב"הוצאות"',
             '3. עקוב אחר הלוח בקרה — הכל מתעדכן אוטומטית!']
    for i, s in enumerate(steps):
        r = 33 + i
        c = mrow(r, r, 2, 10)
        c.value = s
        c.font = mfont(size=12, color=WHITE)
        c.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        c.fill = mfill(DARK_NAVY)

    c = mrow(38, 38, 1, 11)
    c.fill = mfill(GOLD)

    c = mrow(40, 41, 2, 10)
    c.value = 'סיסמת עריכה: smoney2025'
    c.font = mfont(size=10, italic=True, color='888888')
    c.alignment = calign()
    c.fill = mfill(NAVY)

    protect(ws)


# ─── SHEET: לוח בקרה (Dashboard) ──────────────────────────────────────────────

def build_dashboard(ws):
    set_defaults(ws, NAVY, zoom=95, freeze='A3', grid=False)

    ws.column_dimensions['A'].width = 2
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18
    for c in range(8, 16):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws.merge_cells('A1:O1')
    t = ws['A1']
    t.value = 'לוח בקרה  |  לאן הכסף שלי נעלם?'
    t.font = mfont(size=20, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 45

    ws.merge_cells('A2:O2')
    ws['A2'].fill = mfill(GOLD)
    ws.row_dimensions[2].height = 4

    # KPI definitions
    kpis = [
        ('💰 סה"כ הכנסות', "=IFERROR(SUM('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$E$4:$E$503),0)", '27AE60'),
        ('💸 סה"כ הוצאות', "=IFERROR(SUM('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F$4:$F$503),0)", 'E74C3C'),
        ('🏦 חיסכון חודשי', '=B4-D4', '2980B9'),
        ('📊 % חיסכון', '=IFERROR(F4/B4,0)', '8E44AD'),
        ('❓ כסף שנעלם', '=MAX(0,D4-B4)', 'E67E22'),
    ]

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 55
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 10

    cols = [2, 4, 6, 8, 10]
    labels_row = 3
    val_row = 4
    sub_row = 5

    for i, (label, formula, color) in enumerate(kpis):
        col = cols[i]
        cl = get_column_letter(col)
        cl2 = get_column_letter(col + 1)

        ws.merge_cells(f'{cl}{labels_row}:{cl2}{labels_row}')
        lc = ws.cell(labels_row, col)
        lc.value = label
        lc.font = mfont(size=10, bold=True, color=WHITE)
        lc.fill = mfill(color)
        lc.alignment = calign()

        ws.merge_cells(f'{cl}{val_row}:{cl2}{val_row}')
        vc = ws.cell(val_row, col)
        vc.value = formula
        vc.font = mfont(size=22, bold=True, color=NAVY)
        vc.fill = mfill(LIGHT_BLUE)
        vc.alignment = calign()
        vc.border = thick_border()
        if i == 3:
            vc.number_format = '0.0%'
        else:
            vc.number_format = '#,##0 "\u20aa"'

        ws.merge_cells(f'{cl}{sub_row}:{cl2}{sub_row}')
        sc = ws.cell(sub_row, col)
        sc.fill = mfill(LIGHT_BLUE)

    # Savings conditional formatting
    ws.conditional_formatting.add('F4',
        FormulaRule(formula=['F4>0'], fill=mfill('E8F5E9'), font=mfont(size=22, bold=True, color='27AE60')))
    ws.conditional_formatting.add('F4',
        FormulaRule(formula=['F4<0'], fill=mfill('FDECEA'), font=mfont(size=22, bold=True, color='C0392B')))

    ws.row_dimensions[7].height = 10

    ws.merge_cells('B7:H7')
    sh = ws['B7']
    sh.value = 'התפלגות הוצאות לפי קטגוריה'
    sh.font = mfont(size=13, bold=True, color=WHITE)
    sh.fill = mfill(NAVY)
    sh.alignment = calign()

    ws.merge_cells('I7:O7')
    sh2 = ws['I7']
    sh2.value = 'הכנסות מול הוצאות — מגמה שנתית'
    sh2.font = mfont(size=13, bold=True, color=WHITE)
    sh2.fill = mfill(NAVY)
    sh2.alignment = calign()

    # Pie chart from נתונים sheet
    pie = PieChart()
    pie.title = 'הוצאות לפי קטגוריה'
    pie.style = 10
    pie.width = 14
    pie.height = 13
    labels = Reference(ws.parent['נתונים'], min_col=5, min_row=2, max_row=13)
    data = Reference(ws.parent['נתונים'], min_col=6, min_row=1, max_row=13)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = None
    ws.add_chart(pie, 'B8')

    # Bar chart from נתונים sheet
    bar = BarChart()
    bar.type = 'col'
    bar.grouping = 'clustered'
    bar.title = 'מגמה שנתית'
    bar.style = 10
    bar.width = 15
    bar.height = 13
    bar.y_axis.title = 'סכום (₪)'
    data_inc = Reference(ws.parent['נתונים'], min_col=8, min_row=1, max_row=13)
    data_exp = Reference(ws.parent['נתונים'], min_col=9, min_row=1, max_row=13)
    cats = Reference(ws.parent['נתונים'], min_col=7, min_row=2, max_row=13)
    bar.add_data(data_inc, titles_from_data=True)
    bar.add_data(data_exp, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = '27AE60'
    bar.series[1].graphicalProperties.solidFill = 'E74C3C'
    ws.add_chart(bar, 'I8')

    ws.merge_cells('A30:O30')
    ts = ws['A30']
    ts.value = '=TEXT(TODAY(),"עדכון אחרון: DD/MM/YYYY")'
    ts.font = mfont(size=9, italic=True, color='888888')
    ts.alignment = calign()

    protect(ws)


# ─── SHEET: הכנסות (Income) ───────────────────────────────────────────────────

def build_income(ws):
    set_defaults(ws, '27AE60', zoom=100, freeze='A4')

    cols_w = [14, 24, 32, 22, 16]
    for i, w in enumerate(cols_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = 'הכנסות — מעקב מקורות הכנסה'
    t.font = mfont(size=16, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 35

    headers = ['תאריך', 'סוג הכנסה', 'תיאור', 'הערות', 'סכום (₪)']
    for i, h in enumerate(headers, 1):
        header_cell(ws.cell(2, i), h, size=12)
    ws.row_dimensions[2].height = 28

    # Totals row
    ws.merge_cells('A3:D3')
    tc = ws['A3']
    tc.value = 'סה"כ הכנסות'
    tc.font = mfont(size=12, bold=True, color=NAVY)
    tc.fill = mfill(LIGHT_BLUE)
    tc.alignment = ralign()
    tc.border = mborder()
    total = ws['E3']
    total.value = "=IFERROR(SUM(E4:E503),0)"
    total.font = mfont(size=13, bold=True, color='27AE60')
    total.fill = mfill(LIGHT_BLUE)
    total.alignment = calign()
    total.border = thick_border()
    total.number_format = '#,##0 "\u20aa"'
    ws.row_dimensions[3].height = 28

    dv = DataValidation(
        type='list',
        formula1='"משכורת,עבודה עצמאית/פרילנס,בונוס,שכר דירה,דיבידנד/השקעות,מתנה,ירושה,קצבה,אחר"',
        allow_blank=True
    )
    dv.error = 'בחר מהרשימה'
    dv.errorTitle = 'ערך לא תקין'
    ws.add_data_validation(dv)
    dv.sqref = 'B4:B503'

    for row in range(4, 504):
        ws.row_dimensions[row].height = 20
        for col in range(1, 6):
            c = ws.cell(row, col)
            input_cell(c)
            if col == 5:
                c.number_format = '#,##0 "\u20aa"'
            if col == 1:
                c.number_format = 'DD/MM/YYYY'

    ws.conditional_formatting.add(
        'E4:E503',
        ColorScaleRule(start_type='min', start_color='FFFFFF',
                       end_type='max', end_color='27AE60'))

    # Sample income data
    income_data = [
        (date(2025, 1, 1),  'משכורת',              'משכורת ינואר',    '',          12500),
        (date(2025, 1, 15), 'עבודה עצמאית/פרילנס', 'פרויקט עיצוב',    'לקוח A',    2800),
        (date(2025, 2, 1),  'משכורת',              'משכורת פברואר',   '',          12500),
        (date(2025, 2, 20), 'בונוס',               'בונוס רבעוני',    '',          3000),
        (date(2025, 3, 1),  'משכורת',              'משכורת מרץ',      '',          12500),
        (date(2025, 3, 10), 'עבודה עצמאית/פרילנס', 'ייעוץ',           '',          1500),
        (date(2025, 4, 1),  'משכורת',              'משכורת אפריל',    '',          12500),
        (date(2025, 4, 5),  'שכר דירה',            'שכר דירה — חדר',  '',          1800),
        (date(2025, 4, 22), 'עבודה עצמאית/פרילנס', 'פרויקט אתר',      'לקוח B',    3500),
    ]
    for i, row_data in enumerate(income_data, 4):
        ws.cell(i, 1).value = row_data[0]
        ws.cell(i, 2).value = row_data[1]
        ws.cell(i, 3).value = row_data[2]
        ws.cell(i, 4).value = row_data[3]
        ws.cell(i, 5).value = row_data[4]

    tab = Table(displayName='TableIncome', ref='A2:E503')
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    ws.add_table(tab)

    for row in range(4, 504):
        for col in range(1, 6):
            ws.cell(row, col).protection = Protection(locked=False)
    protect(ws)


# ─── SHEET: הוצאות (Expenses) ─────────────────────────────────────────────────

def build_expenses(ws):
    set_defaults(ws, 'E74C3C', zoom=100, freeze='A4')

    cols_w = [14, 22, 22, 30, 20, 16, 18]
    for i, w in enumerate(cols_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:G1')
    t = ws['A1']
    t.value = 'הוצאות — מעקב מלא'
    t.font = mfont(size=16, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 35

    headers = ['תאריך', 'קטגוריה', 'תת-קטגוריה', 'תיאור', 'אמצעי תשלום', 'סכום (₪)', 'הכרחי/לא']
    for i, h in enumerate(headers, 1):
        header_cell(ws.cell(2, i), h, size=12)
    ws.row_dimensions[2].height = 28

    ws.merge_cells('A3:E3')
    tc = ws['A3']
    tc.value = 'סה"כ הוצאות'
    tc.font = mfont(size=12, bold=True, color=NAVY)
    tc.fill = mfill(LIGHT_BLUE)
    tc.alignment = ralign()
    tc.border = mborder()
    total = ws['F3']
    total.value = "=IFERROR(SUM(F4:F503),0)"
    total.font = mfont(size=13, bold=True, color='E74C3C')
    total.fill = mfill(LIGHT_BLUE)
    total.alignment = calign()
    total.border = thick_border()
    total.number_format = '#,##0 "\u20aa"'
    ws.row_dimensions[3].height = 28

    dv_cat = DataValidation(type='list', formula1="'\u05e0\u05ea\u05d5\u05e0\u05d9\u05dd'!$A$1:$A$12", allow_blank=True)
    dv_cat.error = 'בחר מהרשימה'
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = 'B4:B503'

    dv_sub = DataValidation(type='list', formula1="'\u05e0\u05ea\u05d5\u05e0\u05d9\u05dd'!$B$1:$B$40", allow_blank=True)
    ws.add_data_validation(dv_sub)
    dv_sub.sqref = 'C4:C503'

    dv_pay = DataValidation(type='list', formula1='"מזומן,כרטיס אשראי,העברה בנקאית,ביט/פייבוקס,צ\'ק"', allow_blank=True)
    ws.add_data_validation(dv_pay)
    dv_pay.sqref = 'E4:E503'

    dv_ess = DataValidation(type='list', formula1='"הכרחי,לא הכרחי,אופציונלי"', allow_blank=True)
    ws.add_data_validation(dv_ess)
    dv_ess.sqref = 'G4:G503'

    for row in range(4, 504):
        ws.row_dimensions[row].height = 20
        for col in range(1, 8):
            c = ws.cell(row, col)
            input_cell(c)
            if col == 6:
                c.number_format = '#,##0 "\u20aa"'
            if col == 1:
                c.number_format = 'DD/MM/YYYY'

    ws.conditional_formatting.add('A4:G503',
        FormulaRule(formula=['$G4="לא הכרחי"'], fill=mfill('FFF3E0')))
    ws.conditional_formatting.add('F4:F503',
        CellIsRule(operator='greaterThan', formula=['2000'],
                   font=mfont(bold=True, color='C0392B')))
    ws.conditional_formatting.add('F4:F503',
        ColorScaleRule(start_type='min', start_color='FFFFFF',
                       mid_type='percentile', mid_value=50, mid_color='FFD700',
                       end_type='max', end_color='E74C3C'))

    # Sample expense data
    expense_data = [
        (date(2025,1,3),  'דיור',            'שכירות',        'שכירות ינואר',        'העברה בנקאית', 3500, 'הכרחי'),
        (date(2025,1,4),  'דיור',            'חשמל',          'חשמל ינואר',           'העברה בנקאית',  280, 'הכרחי'),
        (date(2025,1,5),  'מזון',            'סופרמרקט',      'שופרסל',               'כרטיס אשראי',   780, 'הכרחי'),
        (date(2025,1,8),  'פנאי ובידור',     'מנויים',        'נטפליקס',              'כרטיס אשראי',    55, 'לא הכרחי'),
        (date(2025,1,9),  'מנויים',          'ספוטיפיי',      'ספוטיפיי',             'כרטיס אשראי',    20, 'לא הכרחי'),
        (date(2025,1,10), 'תחבורה',          'דלק',           'פז',                   'כרטיס אשראי',   310, 'הכרחי'),
        (date(2025,1,12), 'מזון',            'מסעדות',        'ארוחת ערב בחוץ',       'כרטיס אשראי',   240, 'לא הכרחי'),
        (date(2025,1,15), 'בריאות',          'ביטוח בריאות',  'מכבי',                 'העברה בנקאית',  350, 'הכרחי'),
        (date(2025,1,18), 'קניות',           'בגדים',         'זארה',                 'כרטיס אשראי',   450, 'לא הכרחי'),
        (date(2025,1,20), 'מזון',            'קפה',           'קפה גרג\'',            'מזומן',          85, 'לא הכרחי'),
        (date(2025,1,22), 'חינוך',           'קורסים',        'קורס אונליין',         'כרטיס אשראי',  299, 'אופציונלי'),
        (date(2025,1,25), 'תחבורה',          'תחבורה ציבורית','כרטיס חופשי חודשי',   'כרטיס אשראי',  230, 'הכרחי'),
        (date(2025,1,28), 'פנאי ובידור',     'קולנוע',        'סינמה סיטי',           'כרטיס אשראי',  100, 'לא הכרחי'),
        (date(2025,2,3),  'דיור',            'שכירות',        'שכירות פברואר',        'העברה בנקאית', 3500, 'הכרחי'),
        (date(2025,2,5),  'מזון',            'סופרמרקט',      'רמי לוי',              'כרטיס אשראי',  650, 'הכרחי'),
        (date(2025,2,7),  'מזון',            'משלוחים',       'וולט',                 'כרטיס אשראי',  180, 'לא הכרחי'),
        (date(2025,2,10), 'תחבורה',          'דלק',           'סונול',                'כרטיס אשראי',  290, 'הכרחי'),
        (date(2025,2,14), 'מתנות ותרומות',   'מתנה',          'מתנת יום הולדת',       'מזומן',         200, 'אופציונלי'),
        (date(2025,2,18), 'בריאות',          'רופא',          'ביקור רופא משפחה',     'מזומן',          50, 'הכרחי'),
        (date(2025,2,20), 'קניות',           'אלקטרוניקה',   'אוזניות',              'כרטיס אשראי',  490, 'לא הכרחי'),
        (date(2025,3,3),  'דיור',            'שכירות',        'שכירות מרץ',           'העברה בנקאית', 3500, 'הכרחי'),
        (date(2025,3,5),  'מזון',            'סופרמרקט',      'שופרסל',               'כרטיס אשראי',  720, 'הכרחי'),
        (date(2025,3,8),  'ספורט',           'חדר כושר',      'מנוי חדר כושר',        'כרטיס אשראי',  180, 'הכרחי'),
        (date(2025,3,12), 'תחבורה',          'דלק',           'פז',                   'כרטיס אשראי',  340, 'הכרחי'),
        (date(2025,3,15), 'פנאי ובידור',     'חגים וטיולים', 'סוף שבוע בצפון',       'כרטיס אשראי',  850, 'לא הכרחי'),
        (date(2025,3,20), 'חינוך',           'ספרים',         'ספרים מקצועיים',       'כרטיס אשראי',  120, 'אופציונלי'),
        (date(2025,3,25), 'מזון',            'מסעדות',        'ארוחה עם חברים',       'כרטיס אשראי',  320, 'לא הכרחי'),
        (date(2025,4,3),  'דיור',            'שכירות',        'שכירות אפריל',         'העברה בנקאית', 3500, 'הכרחי'),
        (date(2025,4,5),  'דיור',            'אינטרנט',       'HOT',                  'העברה בנקאית',  120, 'הכרחי'),
        (date(2025,4,6),  'מזון',            'סופרמרקט',      'רמי לוי',              'כרטיס אשראי',  690, 'הכרחי'),
        (date(2025,4,10), 'תחבורה',          'אחזקת רכב',    'טיפול רכב',            'מזומן',         580, 'הכרחי'),
        (date(2025,4,12), 'קניות',           'בגדים',         'H&M',                  'כרטיס אשראי',  380, 'לא הכרחי'),
        (date(2025,4,15), 'בריאות',          'תרופות',        'בית מרקחת',            'מזומן',          75, 'הכרחי'),
        (date(2025,4,18), 'מזון',            'קפה',           'ארומה',                'מזומן',          95, 'לא הכרחי'),
        (date(2025,4,20), 'חיסכון והשקעות', 'חיסכון',        'העברה לחיסכון',        'העברה בנקאית', 1000, 'הכרחי'),
        (date(2025,4,21), 'פנאי ובידור',    'בידור',         'הופעה',                'מזומן',         250, 'לא הכרחי'),
    ]
    for i, row_data in enumerate(expense_data, 4):
        ws.cell(i, 1).value = row_data[0]
        ws.cell(i, 2).value = row_data[1]
        ws.cell(i, 3).value = row_data[2]
        ws.cell(i, 4).value = row_data[3]
        ws.cell(i, 5).value = row_data[4]
        ws.cell(i, 6).value = row_data[5]
        ws.cell(i, 7).value = row_data[6]

    tab = Table(displayName='TableExpenses', ref='A2:G503')
    tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium3', showRowStripes=True)
    ws.add_table(tab)

    for row in range(4, 504):
        for col in range(1, 8):
            ws.cell(row, col).protection = Protection(locked=False)
    protect(ws)


# ─── SHEET: תקציב (Budget) ────────────────────────────────────────────────────

def build_budget(ws):
    set_defaults(ws, '8E44AD', zoom=100, freeze='A3')

    cols_w = [28, 18, 18, 16, 16, 20]
    for i, w in enumerate(cols_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = 'תקציב חודשי — מעקב ובקרה'
    t.font = mfont(size=16, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 35

    headers = ['קטגוריה', 'תקציב חודשי', 'בוצע בפועל', 'הפרש', '% ניצול', 'סטטוס']
    for i, h in enumerate(headers, 1):
        header_cell(ws.cell(2, i), h, size=12)
    ws.row_dimensions[2].height = 28

    budget_defaults = {
        'דיור': 3800, 'מזון': 2500, 'תחבורה': 800, 'בריאות': 400,
        'פנאי ובידור': 600, 'קניות': 700, 'חינוך': 300,
        'חיסכון והשקעות': 1500, 'מנויים': 150, 'ילדים': 0,
        'מתנות ותרומות': 200, 'אחר': 300,
    }

    for i, cat in enumerate(CATEGORIES, 3):
        row = i
        ws.row_dimensions[row].height = 24
        c_cat = ws.cell(row, 1)
        c_cat.value = cat
        c_cat.font = mfont(size=11, bold=True, color=NAVY)
        c_cat.fill = mfill(SOFT_GRAY)
        c_cat.alignment = ralign()
        c_cat.border = mborder()

        c_budget = ws.cell(row, 2)
        c_budget.value = budget_defaults.get(cat, 0)
        c_budget.number_format = '#,##0 "\u20aa"'
        c_budget.font = mfont(size=11, color='0000FF')
        input_cell(c_budget)
        c_budget.protection = Protection(locked=False)

        actual_formula = f"=IFERROR(SUMIF('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$B:$B,A{row},'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F:$F),0)"
        c_actual = ws.cell(row, 3)
        c_actual.value = actual_formula
        c_actual.number_format = '#,##0 "\u20aa"'
        calc_cell(c_actual)

        c_diff = ws.cell(row, 4)
        c_diff.value = f'=B{row}-C{row}'
        c_diff.number_format = '#,##0 "\u20aa"'
        calc_cell(c_diff)

        c_pct = ws.cell(row, 5)
        c_pct.value = f'=IFERROR(C{row}/B{row},0)'
        c_pct.number_format = '0.0%'
        calc_cell(c_pct)

        c_status = ws.cell(row, 6)
        c_status.value = f'=IF(E{row}>=1,"חריגה! \u26a0",IF(E{row}>=0.8,"קרוב לגבול \U0001f7e1","בסדר \u2713"))'
        c_status.font = mfont(size=11, bold=True, color=NAVY)
        c_status.alignment = calign()
        c_status.border = mborder()

    # Totals row
    row = 15
    ws.row_dimensions[row].height = 28
    tc = ws.cell(row, 1)
    tc.value = 'סה"כ'
    tc.font = mfont(size=12, bold=True, color=WHITE)
    tc.fill = mfill(NAVY)
    tc.alignment = ralign()
    tc.border = thick_border()
    for col in range(2, 7):
        c = ws.cell(row, col)
        if col == 2:
            c.value = '=SUM(B3:B14)'
        elif col == 3:
            c.value = '=SUM(C3:C14)'
        elif col == 4:
            c.value = '=SUM(D3:D14)'
        elif col == 5:
            c.value = '=IFERROR(C15/B15,0)'
            c.number_format = '0.0%'
        elif col == 6:
            c.value = ''
        if col in (2, 3, 4):
            c.number_format = '#,##0 "\u20aa"'
        c.font = mfont(size=12, bold=True, color=WHITE)
        c.fill = mfill(NAVY)
        c.alignment = calign()
        c.border = thick_border()

    # Traffic light CF
    budget_range = 'E3:E14'
    ws.conditional_formatting.add(budget_range,
        CellIsRule(operator='greaterThanOrEqual', formula=['1'],
                   fill=mfill(RED_ALERT), font=mfont(bold=True, color=WHITE)))
    ws.conditional_formatting.add(budget_range,
        CellIsRule(operator='between', formula=['0.8', '0.9999'],
                   fill=mfill(YELLOW_WARN)))
    ws.conditional_formatting.add(budget_range,
        CellIsRule(operator='lessThan', formula=['0.8'],
                   fill=mfill(GREEN_OK), font=mfont(color=WHITE)))

    ws.conditional_formatting.add('F3:F14',
        FormulaRule(formula=['E3>=1'], fill=mfill(RED_ALERT), font=mfont(bold=True, color=WHITE)))
    ws.conditional_formatting.add('F3:F14',
        FormulaRule(formula=['AND(E3>=0.8,E3<1)'], fill=mfill(YELLOW_WARN)))
    ws.conditional_formatting.add('F3:F14',
        FormulaRule(formula=['E3<0.8'], fill=mfill(GREEN_OK), font=mfont(color=WHITE)))

    ws.conditional_formatting.add('D3:D14',
        FormulaRule(formula=['D3<0'], fill=mfill('FDECEA'), font=mfont(color='C0392B')))
    ws.conditional_formatting.add('D3:D14',
        FormulaRule(formula=['D3>=0'], fill=mfill('E8F5E9'), font=mfont(color='27AE60')))

    ws.conditional_formatting.add('C3:C14',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=5000, color=NAVY))

    # Unlock budget input cells before protecting
    for row in range(3, 15):
        ws.cell(row, 2).protection = Protection(locked=False)
    protect(ws)


# ─── SHEET: סיכום שנתי (Annual) ───────────────────────────────────────────────

def build_annual(ws):
    set_defaults(ws, '2980B9', zoom=100, freeze='A3')

    cols_w = [16, 18, 18, 18, 16]
    for i, w in enumerate(cols_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = 'סיכום שנתי 2025'
    t.font = mfont(size=16, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 35

    headers = ['חודש', 'הכנסות', 'הוצאות', 'חיסכון', '% חיסכון']
    for i, h in enumerate(headers, 1):
        header_cell(ws.cell(2, i), h, size=12)
    ws.row_dimensions[2].height = 28

    months_he = ['ינואר','פברואר','מרץ','אפריל','מאי','יוני',
                 'יולי','אוגוסט','ספטמבר','אוקטובר','נובמבר','דצמבר']

    for m in range(1, 13):
        row = m + 2
        ws.row_dimensions[row].height = 24
        c_month = ws.cell(row, 1)
        c_month.value = months_he[m - 1]
        c_month.font = mfont(size=11, bold=True, color=NAVY)
        c_month.fill = mfill(SOFT_GRAY)
        c_month.alignment = calign()
        c_month.border = mborder()

        inc = (f"=IFERROR(SUMPRODUCT((MONTH('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$A$4:$A$503)={m})"
               f"*(YEAR('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$A$4:$A$503)=2025)"
               f"*'\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$E$4:$E$503),0)")
        exp = (f"=IFERROR(SUMPRODUCT((MONTH('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$A$4:$A$503)={m})"
               f"*(YEAR('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$A$4:$A$503)=2025)"
               f"*'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F$4:$F$503),0)")

        for col, formula in [(2, inc), (3, exp)]:
            c = ws.cell(row, col)
            c.value = formula
            c.number_format = '#,##0 "\u20aa"'
            calc_cell(c)

        c_sav = ws.cell(row, 4)
        c_sav.value = f'=B{row}-C{row}'
        c_sav.number_format = '#,##0 "\u20aa"'
        calc_cell(c_sav)

        c_pct = ws.cell(row, 5)
        c_pct.value = f'=IFERROR(D{row}/B{row},0)'
        c_pct.number_format = '0.0%'
        calc_cell(c_pct)

    # Totals row
    row = 15
    ws.row_dimensions[row].height = 28
    tc = ws.cell(row, 1)
    tc.value = 'סה"כ שנתי'
    tc.font = mfont(size=12, bold=True, color=WHITE)
    tc.fill = mfill(NAVY)
    tc.alignment = calign()
    tc.border = thick_border()
    for col in range(2, 6):
        c = ws.cell(row, col)
        if col in (2, 3, 4):
            c.value = f'=SUM({get_column_letter(col)}3:{get_column_letter(col)}14)'
            c.number_format = '#,##0 "\u20aa"'
        else:
            c.value = '=IFERROR(D15/B15,0)'
            c.number_format = '0.0%'
        c.font = mfont(size=12, bold=True, color=WHITE)
        c.fill = mfill(NAVY)
        c.alignment = calign()
        c.border = thick_border()

    ws.conditional_formatting.add('D3:D14',
        ColorScaleRule(start_type='min', start_color='FF4444',
                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                       end_type='max', end_color='44BB44'))
    ws.conditional_formatting.add('E3:E14',
        ColorScaleRule(start_type='min', start_color='FF4444',
                       mid_type='num', mid_value=0, mid_color='FFFFFF',
                       end_type='max', end_color='44BB44'))
    ws.conditional_formatting.add('B3:B14',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=20000, color='27AE60'))
    ws.conditional_formatting.add('C3:C14',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=20000, color='E74C3C'))

    # Line chart
    lc = LineChart()
    lc.title = 'מגמה שנתית — הכנסות מול הוצאות'
    lc.style = 10
    lc.width = 20
    lc.height = 10
    data_inc = Reference(ws, min_col=2, min_row=2, max_row=14)
    data_exp = Reference(ws, min_col=3, min_row=2, max_row=14)
    cats = Reference(ws, min_col=1, min_row=3, max_row=14)
    lc.add_data(data_inc, titles_from_data=True)
    lc.add_data(data_exp, titles_from_data=True)
    lc.set_categories(cats)
    lc.series[0].graphicalProperties.line.solidFill = '27AE60'
    lc.series[0].graphicalProperties.line.width = 25000
    lc.series[1].graphicalProperties.line.solidFill = 'E74C3C'
    lc.series[1].graphicalProperties.line.width = 25000
    ws.add_chart(lc, 'A17')

    protect(ws)


# ─── SHEET: ניתוח הוצאות (Analysis) ──────────────────────────────────────────

def build_analysis(ws):
    set_defaults(ws, 'E67E22', zoom=100, freeze='A2', grid=False)

    for i, w in enumerate([28, 18, 14, 2, 24, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = 'ניתוח הוצאות — לאן הכסף שלי נעלם?'
    t.font = mfont(size=16, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 35

    # Section A header
    ws.merge_cells('A2:C2')
    h = ws['A2']
    h.value = 'הוצאות לפי קטגוריה'
    h.font = mfont(size=12, bold=True, color=WHITE)
    h.fill = mfill('E67E22')
    h.alignment = calign()
    ws.row_dimensions[2].height = 24

    header_cell(ws['A3'], 'קטגוריה', size=11)
    header_cell(ws['B3'], 'סה"כ הוצאות', size=11)
    header_cell(ws['C3'], '% מסך הכל', size=11)
    ws.row_dimensions[3].height = 24

    for i, cat in enumerate(CATEGORIES, 4):
        row = i
        ws.row_dimensions[row].height = 22
        c = ws.cell(row, 1)
        c.value = cat
        c.font = mfont(size=11, bold=True, color=NAVY)
        c.fill = mfill(SOFT_GRAY if i % 2 == 0 else WHITE)
        c.alignment = ralign()
        c.border = mborder()

        cb = ws.cell(row, 2)
        cb.value = f"=IFERROR(SUMIF('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$B:$B,A{row},'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F:$F),0)"
        cb.number_format = '#,##0 "\u20aa"'
        calc_cell(cb)

        cp = ws.cell(row, 3)
        cp.value = f'=IFERROR(B{row}/SUM($B$4:$B$15),0)'
        cp.number_format = '0.0%'
        calc_cell(cp)

    ws.conditional_formatting.add('B4:B15',
        ColorScaleRule(start_type='min', start_color='FFFFFF',
                       mid_type='percentile', mid_value=50, mid_color='FFD700',
                       end_type='max', end_color='E74C3C'))
    ws.conditional_formatting.add('C4:C15',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color=NAVY))

    # Section B — Top 5
    ws.merge_cells('A17:C17')
    h2 = ws['A17']
    h2.value = 'TOP 5 — הוצאות הגבוהות ביותר'
    h2.font = mfont(size=12, bold=True, color=WHITE)
    h2.fill = mfill(NAVY)
    h2.alignment = calign()
    ws.row_dimensions[17].height = 24

    header_cell(ws['A18'], 'קטגוריה', size=11)
    header_cell(ws['B18'], 'סכום', size=11)
    header_cell(ws['C18'], 'דירוג', size=11)
    ws.row_dimensions[18].height = 24

    for rank in range(1, 6):
        row = 18 + rank
        ws.row_dimensions[row].height = 22
        c_name = ws.cell(row, 1)
        c_name.value = f'=IFERROR(INDEX($A$4:$A$15,MATCH(LARGE($B$4:$B$15,{rank}),$B$4:$B$15,0)),"")'
        c_name.font = mfont(size=11, bold=True, color=NAVY)
        c_name.fill = mfill(LIGHT_BLUE)
        c_name.alignment = ralign()
        c_name.border = mborder()

        c_amt = ws.cell(row, 2)
        c_amt.value = f'=IFERROR(LARGE($B$4:$B$15,{rank}),0)'
        c_amt.number_format = '#,##0 "\u20aa"'
        calc_cell(c_amt)

        c_rank = ws.cell(row, 3)
        c_rank.value = f'#{rank}'
        c_rank.font = mfont(size=13, bold=True, color=GOLD)
        c_rank.fill = mfill(NAVY)
        c_rank.alignment = calign()
        c_rank.border = mborder()

    # Section C — Insights
    ws.merge_cells('A25:C25')
    h3 = ws['A25']
    h3.value = 'תובנות פיננסיות'
    h3.font = mfont(size=12, bold=True, color=WHITE)
    h3.fill = mfill('8E44AD')
    h3.alignment = calign()
    ws.row_dimensions[25].height = 24

    insights = [
        ('סה"כ הוצאות',             "=IFERROR(SUM(B4:B15),0)",                                                   '#,##0 "\u20aa"'),
        ('הוצאות לא הכרחיות',        "=IFERROR(SUMIF('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$G:$G,\"לא הכרחי\",'\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F:$F),0)", '#,##0 "\u20aa"'),
        ('% הוצ׳ לא הכרחיות',        '=IFERROR(B27/B26,0)',                                                       '0.0%'),
        ('פוטנציאל חיסכון (30%)',     '=B27*0.3',                                                                  '#,##0 "\u20aa"'),
        ('כסף שנעלם (הוצ׳ > הכנ׳)', "=MAX(0,SUM('\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'!$F:$F)-SUM('\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'!$E:$E))", '#,##0 "\u20aa"'),
    ]
    for i, (label, formula, fmt) in enumerate(insights, 26):
        row = i
        ws.row_dimensions[row].height = 24
        lc = ws.cell(row, 1)
        lc.value = label
        lc.font = mfont(size=11, bold=True, color=NAVY)
        lc.fill = mfill(LIGHT_BLUE)
        lc.alignment = ralign()
        lc.border = mborder()
        vc = ws.cell(row, 2)
        vc.value = formula
        vc.number_format = fmt
        vc.font = mfont(size=12, bold=True, color=NAVY)
        calc_cell(vc)

    # Pie chart
    pie = PieChart()
    pie.title = 'התפלגות הוצאות'
    pie.style = 10
    pie.width = 14
    pie.height = 14
    labels = Reference(ws, min_col=1, min_row=4, max_row=15)
    data = Reference(ws, min_col=2, min_row=3, max_row=15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, 'E2')

    protect(ws)


# ─── SHEET: הוראות שימוש (Guide) ──────────────────────────────────────────────

def build_guide(ws):
    set_defaults(ws, '95A5A6', zoom=100, grid=False)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20

    ws.merge_cells('A1:C1')
    t = ws['A1']
    t.value = 'הוראות שימוש — לאן הכסף שלי נעלם?'
    t.font = mfont(size=18, bold=True, color=GOLD)
    t.fill = mfill(NAVY)
    t.alignment = calign()
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:C2')
    sub = ws['A2']
    sub.value = 'ברוכים הבאים! הקובץ הזה עוזר לך להבין לאן הכסף שלך הולך — ולקחת שליטה מלאה על הכספים שלך.'
    sub.font = mfont(size=12, italic=True, color=NAVY)
    sub.fill = mfill(LIGHT_BLUE)
    sub.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2, wrap_text=True)
    ws.row_dimensions[2].height = 35

    steps = [
        ('שלב 1 — הגדרת תקציב', [
            '1. לחץ על הלשונית "תקציב" (סגולה)',
            '2. בעמודה "תקציב חודשי" (צהובה) — הזן את הסכום שאתה רוצה להוציא בכל קטגוריה',
            '3. זה הבסיס שלך! הכל מתעדכן אוטומטית.',
        ]),
        ('שלב 2 — הזנת הכנסות', [
            '1. לחץ על הלשונית "הכנסות" (ירוקה)',
            '2. הזן: תאריך, סוג הכנסה (בחר מהרשימה), תיאור, וסכום',
            '3. הסה"כ מתעדכן אוטומטית בשורה 3',
        ]),
        ('שלב 3 — הזנת הוצאות', [
            '1. לחץ על הלשונית "הוצאות" (אדומה)',
            '2. הזן: תאריך, קטגוריה, תת-קטגוריה, תיאור, אמצעי תשלום, וסכום',
            '3. סמן "הכרחי" או "לא הכרחי" — זה חשוב לניתוח!',
        ]),
        ('שלב 4 — לוח הבקרה', [
            '1. לחץ על "לוח בקרה" (כחול כהה) לסקירה מיידית',
            '2. תראה: סה"כ הכנסות, הוצאות, חיסכון, אחוז חיסכון וכסף שנעלם',
            '3. הגרפים מתעדכנים אוטומטית עם כל הזנה',
        ]),
        ('שלב 5 — ניתוח "לאן הכסף נעלם"', [
            '1. לחץ על "ניתוח הוצאות" (כתום)',
            '2. תראה גרף עוגה ורשימת קטגוריות לפי סדר הוצאה',
            '3. שים לב לשדה "פוטנציאל חיסכון" — כמה אפשר לחסוך בקלות!',
        ]),
        ('שלב 6 — מעקב תקציב', [
            '1. לחץ על "תקציב" לראות את מצב התקציב שלך',
            '2. ירוק = בסדר | צהוב = קרוב לגבול | אדום = חריגה!',
            '3. עדכן את יעדי התקציב בכל חודש לפי הצורך',
        ]),
        ('שלב 7 — סיכום שנתי', [
            '1. לחץ על "סיכום שנתי" (כחול) לראות את כל 12 החודשים',
            '2. הגרף הקווי מראה את המגמה לאורך השנה',
            '3. השתמש בזה לתכנון השנה הבאה',
        ]),
        ('טיפים חשובים', [
            '⭐ שמור את הקובץ בשם חדש כל חודש (למשל: חשבון_ינואר_2025.xlsx)',
            '⭐ הזן הוצאות בזמן אמת — לא בסוף החודש',
            '⭐ סיסמת עריכה אם צריך: smoney2025',
            '⭐ תאי צהובים = מקום להזנה | תאי כחולים = חישוב אוטומטי',
        ]),
    ]

    current_row = 4
    for step_title, step_items in steps:
        ws.row_dimensions[current_row].height = 30
        ws.merge_cells(f'A{current_row}:C{current_row}')
        sh = ws.cell(current_row, 1)
        sh.value = step_title
        sh.font = mfont(size=13, bold=True, color=WHITE)
        sh.fill = mfill(NAVY)
        sh.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        current_row += 1

        for item in step_items:
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(f'A{current_row}:C{current_row}')
            ic = ws.cell(current_row, 1)
            ic.value = item
            ic.font = mfont(size=11, color=NAVY)
            ic.fill = mfill(SOFT_GRAY if current_row % 2 == 0 else WHITE)
            ic.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2, wrap_text=True)
            current_row += 1

        current_row += 1  # spacer

    # Color legend
    ws.row_dimensions[current_row].height = 30
    ws.merge_cells(f'A{current_row}:C{current_row}')
    leg = ws.cell(current_row, 1)
    leg.value = 'מקרא צבעים'
    leg.font = mfont(size=13, bold=True, color=WHITE)
    leg.fill = mfill(GOLD)
    leg.alignment = calign()
    current_row += 1

    legend_items = [
        (LIGHT_YELLOW, NAVY, 'תא צהוב = הזנה ידנית שלך'),
        (LIGHT_BLUE, NAVY, 'תא כחול = חישוב אוטומטי — אל תשנה!'),
        (GREEN_OK, WHITE, 'ירוק = בתקציב / חיסכון חיובי'),
        (YELLOW_WARN, NAVY, 'צהוב = קרוב לגבול התקציב (80–99%)'),
        (RED_ALERT, WHITE, 'אדום = חריגה מהתקציב!'),
    ]
    for bg, fg, text in legend_items:
        ws.row_dimensions[current_row].height = 24
        ws.merge_cells(f'A{current_row}:C{current_row}')
        lc = ws.cell(current_row, 1)
        lc.value = text
        lc.font = mfont(size=11, bold=True, color=fg)
        lc.fill = mfill(bg)
        lc.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        current_row += 1


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    OUTPUT_DIR = os.path.expanduser('~/Downloads/Smoney')
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, '\u05dc\u05d0\u05df_\u05d4\u05db\u05e1\u05e3_\u05e9\u05dc\u05d9_\u05e0\u05e2\u05dc\u05dd.xlsx')
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # Rename default and create all sheets
    wb.active.title = '\u05d3\u05e3 \u05e4\u05ea\u05d9\u05d7\u05d4'
    sheet_names = ['\u05dc\u05d5\u05d7 \u05d1\u05e7\u05e8\u05d4',
                   '\u05d4\u05db\u05e0\u05e1\u05d5\u05ea',
                   '\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea',
                   '\u05ea\u05e7\u05e6\u05d5\u05d1',
                   '\u05e1\u05d9\u05db\u05d5\u05dd \u05e9\u05e0\u05ea\u05d9',
                   '\u05e0\u05d9\u05ea\u05d5\u05d7 \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea',
                   '\u05d4\u05d5\u05e8\u05d0\u05d5\u05ea \u05e9\u05d9\u05de\u05d5\u05e9',
                   '\u05e0\u05ea\u05d5\u05e0\u05d9\u05dd']
    for name in sheet_names:
        wb.create_sheet(name)

    # Build sheets (data sheet first for chart references)
    build_data_sheet(wb['\u05e0\u05ea\u05d5\u05e0\u05d9\u05dd'])
    build_income(wb['\u05d4\u05db\u05e0\u05e1\u05d5\u05ea'])
    build_expenses(wb['\u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'])
    build_budget(wb['\u05ea\u05e7\u05e6\u05d5\u05d1'])
    build_annual(wb['\u05e1\u05d9\u05db\u05d5\u05dd \u05e9\u05e0\u05ea\u05d9'])
    build_analysis(wb['\u05e0\u05d9\u05ea\u05d5\u05d7 \u05d4\u05d5\u05e6\u05d0\u05d5\u05ea'])
    build_dashboard(wb['\u05dc\u05d5\u05d7 \u05d1\u05e7\u05e8\u05d4'])
    build_cover(wb['\u05d3\u05e3 \u05e4\u05ea\u05d9\u05d7\u05d4'])
    build_guide(wb['\u05d4\u05d5\u05e8\u05d0\u05d5\u05ea \u05e9\u05d9\u05de\u05d5\u05e9'])

    # Hide data sheet
    wb['\u05e0\u05ea\u05d5\u05e0\u05d9\u05dd'].sheet_state = 'hidden'

    # Set active sheet to dashboard
    wb.active = wb['\u05dc\u05d5\u05d7 \u05d1\u05e7\u05e8\u05d4']

    wb.save(OUTPUT_FILE)
    print(f'Saved: {OUTPUT_FILE}')


if __name__ == '__main__':
    main()
